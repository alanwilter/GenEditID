import sqlalchemy
import openpyxl
import sys

from dnascissors.config import cfg
from dnascissors.model import *
from openpyxl.workbook import *

from datetime import datetime
from excelloader import ExcelLoader


class LayoutLoader(ExcelLoader):
    
    columnSeparator = ','
    
    projectsFile = None
    targetsFile = None
    guidesFile = None
    layoutFile = None
    
    def toStrand(self, cell, rowNumber):
        str = self.get_string(cell)
        
        if not str:
            return None
        
        if str in ['+', 'positive', 'p', 'forward', 'f']:
            return 'forward'
        
        if str in ['-', 'negative', 'n', 'reverse', 'r']:
            return 'reverse'
        
        raise ValueError('Strand must be "forward" or "reverse", not "%s" (row %d)' % (str, rowNumber))

    
    def loadAll(self, session, wbFile):
        
        workbook = openpyxl.load_workbook(wbFile, read_only=True, data_only=True)
        try:
            self.loadProjects(session, workbook)
            session.commit()
            self.loadTargets(session, workbook)
            session.commit()
            self.loadGuides(session, workbook)
            session.commit()
            self.loadAmpliconSelection(session, workbook)
            session.commit()
            self.loadExperimentLayout(session, workbook)
            session.commit()
        finally:
            workbook.close()


    def loadProjects(self, session, workbook):
        
        sheet = None
        
        try:
            sheet = workbook['Project']
        except KeyError as e:
            raise Exception("There is no 'Project' sheet in the work book.")
        
        rowNumber = 1
        
        for row in sheet.iter_rows(min_row=2):
            ++rowNumber
            
            projectId = self.get_string(row[0])
            
            if not projectId:
                raise Exception('Project identifier is required on row %d' % rowNumber)
            
            project = session.query(Project).filter(Project.geid == projectId).first()
            
            if project:
                print("Already have a project %s (%s)." % (project.geid, project.name))
                return
            
            project = Project()
            project.geid = projectId
            project.name = self.get_string(row[1])
            project.scientist = self.get_string(row[2])
            project.group = self.get_string(row[4])
            project.group_leader = self.get_string(row[5])
            project.start_date = self.get_date(row[6])
            project.description = self.get_string(row[7], 1024)
            
            session.add(project)
            
            print('Created project %s' % project.name)

                
    def loadTargets(self, session, workbook):
            
        sheet = None
        
        try:
            sheet = workbook['Target']
        except KeyError as e:
            raise Exception("There is no 'Target' sheet in the work book.")
        
        rowNumber = 1
        
        for row in sheet.iter_rows(min_row=2):
            ++rowNumber
            
            projectId = self.get_string(row[0])
            
            if not projectId:
                raise Exception('Project identifier is required on row %d' % rowNumber)
            
            project = session.query(Project).filter(Project.geid == projectId).first()
            
            if not project:
                raise Exception('Project "%s" does not exist (row %d' % (projectId, rowNumber))
            
            targetName = self.get_string(row[1])
            
            if not targetName:
                raise Exception('Target name is required on row %d' % rowNumber)
            
            target = session.query(Target).filter(Target.name == targetName).first()
            
            if target:
                print("Already have a target called %s. It's from the project %s." % (target.name, target.project.name))
                return
            
            target = Target(project=project)
            target.name = targetName
            target.species = self.get_string(row[2])
            target.assembly = self.get_string(row[3])
            target.gene_id = self.get_string(row[4])
            target.chromosome = self.get_string(row[5])
            target.start = self.get_int(row[6])
            target.end = self.get_int(row[7])
            target.strand = self.toStrand(row[8], rowNumber)
            target.description = self.get_string(row[9], 1024)
            
            session.add(target)
            
            print('Created target %s' % target.name)

    
    def loadGuides(self, session, workbook):
        sheet = None
        
        try:
            sheet = workbook['Guide']
        except KeyError as e:
            raise Exception("There is no 'Guide' sheet in the work book.")
        
        rowNumber = 1
        
        for row in sheet.iter_rows(min_row=2):
            ++rowNumber
                
            targetName = self.get_string(row[0])
            
            if not targetName:
                raise Exception('Target name is required on row %d' % rowNumber)
            
            target = session.query(Target).filter(Target.name == targetName).first()
            
            if not target:
                raise Exception('Target "%s" does not exist (row %d)' % (targetName, rowNumber))

            guideName = self.get_string(row[1])
            
            if not guideName:
                raise Exception('Guide name is required on row %d' % rowNumber)
            
            guide = Guide(target=target)
            guide.name = guideName
            guide.guide_sequence = self.get_string(row[2])
            guide.pam_sequence = self.get_string(row[3])
            guide.activity = self.get_int(row[4])
            guide.exon = self.get_int(row[5])
            guide.nuclease = self.get_string(row[6])
            
            session.add(guide)
            
            print('Created guide %s' % guide.name)


    def loadAmpliconSelection(self, session, workbook):
        
        sheet = None
        
        try:
            sheet = workbook['AmpliconSelection']
        except KeyError as e:
            raise Exception("There is no 'AmpliconSelection' sheet in the work book.")
        
        rowNumber = 1
        previous_strand = None
            
        for row in sheet.iter_rows(min_row=2):
            ++rowNumber
                
            # Need the guide first.
            
            guideName = self.get_string(row[0])
            
            if not guideName:
                raise Exception('Guide name is required on row %d' % rowNumber)
            
            guide = session.query(Guide).filter(Guide.name == guideName).first()
            
            if not guide:
                raise Exception('Guide "%s" does not exist (row %d)' % (projectId, rowNumber))
            
            
            # Find or create the amplicon
            
            ampliconName = self.get_string(row[5])
            
            if not ampliconName:
                raise Exception('Amplicon name is required on row %d' % rowNumber)

            amplicon = session.query(Amplicon).filter(Amplicon.name == ampliconName).first()
            
            if not amplicon:
                
                onTarget = self.get_string(row[6])
                
                amplicon = Amplicon()
                amplicon.name = ampliconName
                amplicon.is_on_target = onTarget and 'true' == onTarget.casefold()
                amplicon.dna_feature = self.get_string(row[7])
                amplicon.chromosome = self.get_string(row[8])
                
                session.add(amplicon);

                print('Created amplicon %s' % amplicon.name)
            
            
            # Find or create the amplicon selection
            
            guideLocation = self.get_int(row[2])
            
            strand = self.toStrand(row[3], rowNumber)
            if strand:
                previous_strand = strand
            else:
                if previous_strand:
                    strand = previous_strand
                else:
                    raise Exception("Have no guide strand on row %d" % rowNumber)
            
            
            selection = session.query(AmpliconSelection).filter(AmpliconSelection.amplicon == amplicon)\
                                                        .filter(AmpliconSelection.guide_location == guideLocation)\
                                                        .filter(AmpliconSelection.guide_strand == strand).first()
            
            if not selection:
                
                selection = AmpliconSelection(guide=guide, amplicon=amplicon)
                selection.experiment_type = self.get_string(row[1])
                selection.guide_location = guideLocation
                selection.guide_strand = strand
                
                scoreStr = self.get_string(row[4])
                if scoreStr != None and scoreStr != 'NA':
                    selection.score = int(scoreStr)
                
                session.add(selection)
                
                print('Created amplicon selection of amplicon %s at %d' % (amplicon.name, selection.guide_location))

            geid = self.get_string(row[9])
            
            if not geid:
                raise Exception('Primer GEID is required on row %d' % rowNumber)
            
            primer = session.query(Primer).filter(Primer.geid == geid).first()
            
            if not primer:
                primer = Primer()
                primer.geid = self.get_string(row[9])
                primer.sequence = self.get_string(row[10])
                primer.strand = self.toStrand(row[11], rowNumber)
                primer.start = self.get_int(row[12])
                primer.end = self.get_int(row[13])
                primer.description = self.get_string(row[14], 1024)
                
                session.add(primer)
            
                print('Created primer %s' % primer.geid)
                
            
            primer.amplicons.append(amplicon)
            
            print('Linked amplicon %s with primer %s' % (amplicon.name, primer.geid))


    def loadExperimentLayout(self, session, workbook):
        
        sheet = None
        
        try:
            sheet = workbook['ExperimentLayout']
        except KeyError as e:
            raise Exception("There is no 'ExperimentLayout' sheet in the work book.")
        
        rowNumber = 1
        plate = None
            
        for row in sheet.iter_rows(min_row=2):
            ++rowNumber
                
            # Need to find project.
            
            projectId = self.get_string(row[0])
            
            if not projectId:
                raise Exception('Project identifier is required on row %d' % rowNumber)
            
            project = session.query(Project).filter(Project.geid == projectId).first()
            
            if not project:
                raise Exception('Project "%s" does not exist (row %d)' % (projectId, rowNumber))

            # May also need to find a guide.
            
            guide = None
            
            guideName = self.get_string(row[5])
            
            if guideName:
                guide = session.query(Guide).filter(Guide.name == guideName).first()
                
                if not guide:
                    raise("There is no guide with the name %s" % guideName)

            # Experiment layout

            layoutGeid = self.get_string(row[1])
            
            if not layoutGeid:
                raise Exception('Experiment layout GEID is required on row %d' % rowNumber)
            
            layout = session.query(ExperimentLayout).filter(ExperimentLayout.geid == layoutGeid).first()
            
            if not layout:
                layout = ExperimentLayout(project=project)
                layout.geid = layoutGeid
                
                session.add(layout)
                
                print('Created experiment layout %s in project %s' % (layout.geid, project.geid))

            # Cell row
            
            celllineName = self.get_string(row[3])
            
            cellline = None
            
            if celllineName:
                cellline = session.query(CellLine).filter(CellLine.name == celllineName).first()
                
                if not cellline:
                    cellline = CellLine(name=celllineName)
                    
                    session.add(cellline)
                    
                    print('Created cell row %s' % cellline.name)
            
            # Clone.
            
            cloneName = self.get_string(row[4])
            
            clone = None
            
            if cloneName:
                
                if not cellline:
                    raise Exception('Cannot have a clone without a cell row on row %d' % rowNumber)
                
                clone = session.query(Clone).filter(Clone.cell_line == cellline).filter(Clone.name == cloneName).first()
                
                if not clone:
                    clone = Clone(cell_line=cellline)
                    clone.name = cloneName
                    
                    session.add(clone)
                    
                    print('Created clone %s from cell row %s' % (clone.name, clone.cell_line.name))

            # Plate
            
            plate = session.query(Plate).filter(Plate.geid == layoutGeid).first()
            
            if not plate:
                plate = Plate(experiment_layout=layout)
                plate.geid = layout.geid
                
                session.add(plate)
                
                print('Created plate %s' % plate.geid)

            
            # Well content
            
            content = None
            
            if clone:
                
                contentType = self.get_string(row[8])
                if contentType:
                    contentType = contentType.lower()
                    
                    if contentType in ['wt', 'wildtype', 'wild-type']:
                        contentType = 'WT'
                    elif contentType in ['ko', 'knockout', 'knock-out']:
                        contentType = 'KO'
                    elif contentType in ['bg', 'background']:
                        contentType = 'BG'
                    elif contentType in ['nm', 'normalisation', 'normaliser']:
                        contentType = 'NM'
                    elif contentType in ['sm', 'sample']:
                        contentType = 'SM'
                    elif contentType == 'empty':
                        contentType = None
                    else:
                        raise Exception("Well content type not recognised: %s" % contentType)
                
                
                content = WellContent(clone=clone)
                
                if guide:
                    content.guides.append(guide)
                
                controlFlag = self.get_string(row[7])
                
                content.replicate_group = self.get_int(row[6])
                content.is_control = controlFlag and 'true' == controlFlag.casefold()
                content.content_type = contentType
                
                session.add(content)
                
                print("Created well content for clone %s" % content.clone.name)

            
            # Well
            
            wellPos = self.get_string(row[2])
            
            if not wellPos:
                raise Exception('Well position is required on row %d' % rowNumber)
            
            well = Well(experiment_layout=layout)
            well.row = wellPos[0]
            well.column = int(wellPos[1:])
            well.well_content = content
            
            session.add(well)
            
            print('Created well %s%d in layout %s' % (well.row, well.column, well.experiment_layout.geid))


if len(sys.argv) < 2:
    print("Need the layout Excel file.", file=sys.stderr)
    sys.exit(1)

engine = sqlalchemy.create_engine(cfg['DATABASE_URI'])

Base.metadata.bind = engine

DBSession = sqlalchemy.orm.sessionmaker(bind=engine)

session = DBSession()

"""
deleteCount = session.query(CellLine).delete()
print('Deleted %d cell lines' % deleteCount)

deleteCount = session.query(Primer).delete()
print('Deleted %d primers' % deleteCount)

deleteCount = session.query(Amplicon).delete()
print('Deleted %d amplicons' % deleteCount)

deleteCount = session.query(Project).delete()
print('Deleted %d projects' % deleteCount)
"""

loader = LayoutLoader()

try:
    loader.loadAll(session, sys.argv[1])
finally:
    session.close()
