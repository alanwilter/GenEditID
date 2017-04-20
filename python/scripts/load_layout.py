import sqlalchemy
import openpyxl
import logging

from dnascissors.config import cfg
from dnascissors.model import Base, Project, Target, Guide, Amplicon, AmpliconSelection, Primer, CellLine, Clone, ExperimentLayout, Plate, WellContent, Well

from excelloader import ExcelLoader

# logging configuration
import log as logger


class LayoutLoader(ExcelLoader):

    columnSeparator = ','

    def __init__(self, session, workbook_file):
        self.log = logging.getLogger('dnascissors')
        self.session = session
        self.workbook = openpyxl.load_workbook(workbook_file, read_only=True, data_only=True)

    def toStrand(self, cell, rowNumber):
        str = self.get_string(cell)

        if not str:
            return None

        if str in ['+', 'positive', 'p', 'forward', 'f']:
            return 'forward'

        if str in ['-', 'negative', 'n', 'reverse', 'r']:
            return 'reverse'

        raise ValueError('Strand must be "forward" or "reverse", not "{:s}" (row {:d})'.format(str, rowNumber))

    def loadAll(self):
        try:
            self.loadProjects()
            self.session.commit()
            self.loadTargets()
            self.session.commit()
            self.loadGuides()
            self.session.commit()
            self.loadAmpliconSelection()
            self.session.commit()
            self.loadExperimentLayout()
            self.session.commit()
        finally:
            self.workbook.close()

    def loadProjects(self):

        sheet = None

        try:
            sheet = self.workbook['Project']
        except KeyError as e:
            raise Exception("There is no 'Project' sheet in the work book.", e)

        rowNumber = 1

        for row in sheet.iter_rows(min_row=2):
            ++rowNumber

            projectId = self.get_string(row[0])

            if not projectId:
                raise Exception('Project identifier is required on row {:d}'.format(rowNumber))

            project = self.session.query(Project).filter(Project.geid == projectId).first()

            if project:
                self.log.info("Already have a project {:s} ({:s})".format(project.geid, project.name))
                return

            project = Project()
            project.geid = projectId
            project.name = self.get_string(row[1])
            project.scientist = self.get_string(row[2])
            project.group = self.get_string(row[4])
            project.group_leader = self.get_string(row[5])
            project.start_date = self.get_date(row[6])
            project.description = self.get_string(row[7], 1024)

            self.session.add(project)

            self.log.info('Created project {:s}'.format(project.name))

    def loadTargets(self):

        sheet = None

        try:
            sheet = self.workbook['Target']
        except KeyError as e:
            raise Exception("There is no 'Target' sheet in the work book.", e)

        rowNumber = 1

        for row in sheet.iter_rows(min_row=2):
            ++rowNumber

            projectId = self.get_string(row[0])

            if not projectId:
                raise Exception('Project identifier is required on row {:d}'.format(rowNumber))

            project = self.session.query(Project).filter(Project.geid == projectId).first()

            if not project:
                raise Exception('Project "{:s}" does not exist (row {:d})'.format(projectId, rowNumber))

            targetName = self.get_string(row[1])

            if not targetName:
                raise Exception('Target name is required on row {:d}'.format(rowNumber))

            target = self.session.query(Target).filter(Target.name == targetName).first()

            if target:
                self.log.info("Already have a target called {:s}. It's from the project {:s}.".format(target.name, target.project.name))
                return

            target = Target(project=project)
            target.name = targetName
            target.species = self.get_string(row[2])
            target.assembly = self.get_string(row[3])
            target.gene_id = self.get_string(row[4])
            target.chromosome = self.get_string(row[5])
            target.start = self.get_int(row[6])
            target.end = self.get_int(row[7])
            target.strand = self.toStrand(row[8], rowNumber)
            target.description = self.get_string(row[9], 1024)

            self.session.add(target)

            self.log.info('Created target {:s}'.format(target.name))

    def loadGuides(self):
        sheet = None

        try:
            sheet = self.workbook['Guide']
        except KeyError as e:
            raise Exception("There is no 'Guide' sheet in the work book.", e)

        rowNumber = 1

        for row in sheet.iter_rows(min_row=2):
            ++rowNumber

            targetName = self.get_string(row[0])

            if not targetName:
                raise Exception('Target name is required on row {:d}'.format(rowNumber))

            target = self.session.query(Target).filter(Target.name == targetName).first()

            if not target:
                raise Exception('Target "{:s}" does not exist (row {:d})'.format(targetName, rowNumber))

            guideName = self.get_string(row[1])

            if not guideName:
                raise Exception('Guide name is required on row {:d}'.format(rowNumber))

            guide = Guide(target=target)
            guide.name = guideName
            guide.guide_sequence = self.get_string(row[2])
            guide.pam_sequence = self.get_string(row[3])
            guide.activity = self.get_int(row[4])
            guide.exon = self.get_int(row[5])
            guide.nuclease = self.get_string(row[6])

            self.session.add(guide)

            self.log.info('Created guide {:s}'.format(guide.name))

    def loadAmpliconSelection(self):

        sheet = None

        try:
            sheet = self.workbook['AmpliconSelection']
        except KeyError as e:
            raise Exception("There is no 'AmpliconSelection' sheet in the work book.")

        rowNumber = 1
        previous_strand = None

        for row in sheet.iter_rows(min_row=2):
            ++rowNumber

            # Need the guide first.

            guideName = self.get_string(row[0])

            if not guideName:
                raise Exception('Guide name is required on row {:d}'.format(rowNumber))

            guide = self.session.query(Guide).filter(Guide.name == guideName).first()

            if not guide:
                raise Exception('Guide "{:s}" does not exist (row {:d})'.format(projectId, rowNumber))

            # Find or create the amplicon
            ampliconName = self.get_string(row[5])

            if not ampliconName:
                raise Exception('Amplicon name is required on row {:d}'.format(rowNumber))

            amplicon = self.session.query(Amplicon).filter(Amplicon.name == ampliconName).first()

            if not amplicon:

                onTarget = self.get_string(row[6])

                amplicon = Amplicon()
                amplicon.name = ampliconName
                amplicon.is_on_target = onTarget and 'true' == onTarget.casefold()
                amplicon.dna_feature = self.get_string(row[7])
                amplicon.chromosome = self.get_string(row[8])

                self.session.add(amplicon)

                self.log.info('Created amplicon {:s}'.format(amplicon.name))

            # Find or create the amplicon selection

            guideLocation = self.get_int(row[2])

            strand = self.toStrand(row[3], rowNumber)
            if strand:
                previous_strand = strand
            else:
                if previous_strand:
                    strand = previous_strand
                else:
                    raise Exception("Have no guide strand on row {:d}".format(rowNumber))

            selection = self.session.query(AmpliconSelection).filter(AmpliconSelection.amplicon == amplicon)\
                                                        .filter(AmpliconSelection.guide_location == guideLocation)\
                                                        .filter(AmpliconSelection.guide_strand == strand).first()

            if not selection:

                selection = AmpliconSelection(guide=guide, amplicon=amplicon)
                selection.experiment_type = self.get_string(row[1])
                selection.guide_location = guideLocation
                selection.guide_strand = strand

                scoreStr = self.get_string(row[4])
                if scoreStr is not None and scoreStr != 'NA':
                    selection.score = int(scoreStr)

                self.session.add(selection)

                self.log.info('Created amplicon selection of amplicon {:s} at {:d}'.format(amplicon.name, selection.guide_location))

            geid = self.get_string(row[9])

            if not geid:
                raise Exception('Primer GEID is required on row {:d}'.format(rowNumber))

            primer = self.session.query(Primer).filter(Primer.geid == geid).first()

            if not primer:
                primer = Primer()
                primer.geid = self.get_string(row[9])
                primer.sequence = self.get_string(row[10])
                primer.strand = self.toStrand(row[11], rowNumber)
                primer.start = self.get_int(row[12])
                primer.end = self.get_int(row[13])
                primer.description = self.get_string(row[14], 1024)

                self.session.add(primer)

                self.log.info('Created primer {:s}'.format(primer.geid))

            primer.amplicons.append(amplicon)

            self.log.info('Linked amplicon {:s} with primer {:s}'.format(amplicon.name, primer.geid))

    def loadExperimentLayout(self):

        sheet = None

        try:
            sheet = self.workbook['ExperimentLayout']
        except KeyError as e:
            raise Exception("There is no 'ExperimentLayout' sheet in the work book.")

        rowNumber = 1
        plate = None

        for row in sheet.iter_rows(min_row=2):
            ++rowNumber

            # Need to find project.

            projectId = self.get_string(row[0])

            if not projectId:
                raise Exception('Project identifier is required on row {:d}'.format(rowNumber))

            project = self.session.query(Project).filter(Project.geid == projectId).first()

            if not project:
                raise Exception('Project "{:s}" does not exist (row {:d})'.format(projectId, rowNumber))

            # May also need to find a guide.

            guide = None

            guideName = self.get_string(row[5])

            if guideName:
                guide = self.session.query(Guide).filter(Guide.name == guideName).first()

                if not guide:
                    raise("There is no guide with the name {:s}".format(guideName))

            # Experiment layout

            layoutGeid = self.get_string(row[1])

            if not layoutGeid:
                raise Exception('Experiment layout GEID is required on row {:d}'.format(rowNumber))

            layout = self.session.query(ExperimentLayout).filter(ExperimentLayout.geid == layoutGeid).first()

            if not layout:
                layout = ExperimentLayout(project=project)
                layout.geid = layoutGeid

                self.session.add(layout)

                self.log.info('Created experiment layout {:s} in project {:s}'.format(layout.geid, project.geid))

            # Cell row

            celllineName = self.get_string(row[3])

            cellline = None

            if celllineName:
                cellline = self.session.query(CellLine).filter(CellLine.name == celllineName).first()

                if not cellline:
                    cellline = CellLine(name=celllineName)

                    self.session.add(cellline)

                    self.log.info('Created cell row {:s}'.format(cellline.name))

            # Clone

            cloneName = self.get_string(row[4])

            clone = None

            if cloneName:

                if not cellline:
                    raise Exception('Cannot have a clone without a cell row on row {:d}'.format(rowNumber))

                clone = self.session.query(Clone).filter(Clone.cell_line == cellline).filter(Clone.name == cloneName).first()

                if not clone:
                    clone = Clone(cell_line=cellline)
                    clone.name = cloneName

                    self.session.add(clone)

                    self.log.info('Created clone {:s} from cell row {:s}'.format(clone.name, clone.cell_line.name))

            # Plate

            plate = self.session.query(Plate).filter(Plate.geid == layoutGeid).first()

            if not plate:
                plate = Plate(experiment_layout=layout)
                plate.geid = layout.geid

                self.session.add(plate)

                self.log.info('Created plate {:s}'.format(plate.geid))

            # Well content

            content = None

            if clone:

                contentType = self.get_string(row[8])
                if contentType:
                    contentType = contentType.lower()

                    if contentType in ['wt', 'wildtype', 'wild-type']:
                        contentType = 'wild-type'
                    elif contentType in ['ko', 'knockout', 'knock-out']:
                        contentType = 'knock-out'
                    elif contentType in ['bg', 'background']:
                        contentType = 'background'
                    elif contentType in ['nm', 'normalisation', 'normaliser']:
                        contentType = 'normalisation'
                    elif contentType in ['sm', 'sample']:
                        contentType = 'sample'
                    elif contentType == 'empty':
                        contentType = None
                    else:
                        raise Exception("Well content type not recognised: {:s}".format(contentType))

                content = WellContent(clone=clone)

                if guide:
                    content.guides.append(guide)

                controlFlag = self.get_string(row[7])

                content.replicate_group = self.get_int(row[6])
                content.is_control = controlFlag and 'true' == controlFlag.casefold()
                content.content_type = contentType

                self.session.add(content)

                self.log.info("Created well content for clone {:s}".format(content.clone.name))

            # Well
            wellPos = self.get_string(row[2])

            if not wellPos:
                raise Exception('Well position is required on row {:d}'.format(rowNumber))

            well = Well(experiment_layout=layout)
            well.row = wellPos[0]
            well.column = int(wellPos[1:])
            well.well_content = content

            self.session.add(well)

            self.log.info('Created well {:s}{:d} in layout {:s}'.format(well.row, well.column, well.experiment_layout.geid))


def main():
    import argparse
    import os
    parser = argparse.ArgumentParser()
    parser.add_argument("--layout", dest="file_layout", action="store", help="The file layout e.g. '20170118_GEP00001.xlsx'", required=True)
    parser.add_argument("--clean", dest="clean_db", action="store_true", default=False, help="Clean database before loading?")
    options = parser.parse_args()

    log = logger.get_custom_logger(os.path.join(os.path.dirname(__file__), 'load_layout.log'))

    engine = sqlalchemy.create_engine(cfg['DATABASE_URI'])
    Base.metadata.bind = engine
    DBSession = sqlalchemy.orm.sessionmaker(bind=engine)
    session = DBSession()

    if options.clean_db:
        deleteCount = session.query(CellLine).delete()
        log.info('Deleted {:d} cell lines'.format(deleteCount))

        deleteCount = session.query(Primer).delete()
        log.info('Deleted {:d} primers'.format(deleteCount))

        deleteCount = session.query(Amplicon).delete()
        log.info('Deleted {:d} amplicons'.format(deleteCount))

        deleteCount = session.query(Project).delete()
        log.info('Deleted {:d} projects'.format(deleteCount))

    loader = LayoutLoader(session, options.file_layout)

    try:
        loader.loadAll()
    except Exception as e:
        log.exception(e)
    finally:
        session.close()

if __name__ == '__main__':
    main()
